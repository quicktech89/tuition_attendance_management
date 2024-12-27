from flask import Flask, request, render_template, send_from_directory, redirect, url_for
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime

app = Flask(__name__)

# Ensure the static folder exists
if not os.path.exists('static'):
    os.makedirs('static')

# File path for the attendance Excel sheet
file_path = os.path.join('static', 'attendance.xlsx')

# Variable to track if the reset has already occurred
reset_done = False

# Function to format the Excel file
def format_excel():
    """Apply formatting to the Excel sheet."""
    wb = load_workbook(file_path)
    ws = wb.active

    # Add a watermark text in the first row
    if ws.cell(row=1, column=1).value != "Shared by Pratyush":
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws.cell(row=1, column=1).value = "Shared by Pratyush"
        ws.cell(row=1, column=1).font = Font(size=14, italic=True, bold=True, color="808080")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # Format headers
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col in range(1, 3):  # Assuming 2 columns: Date and Reason
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Format the date column to be colorful
    for row in range(3, ws.max_row + 1):
        date_cell = ws.cell(row=row, column=1)
        date_cell.font = Font(color="FF0000")  # Red color for dates
        date_cell.alignment = Alignment(horizontal="center")

    # Adjust column width for the date to make sure it's visible
    ws.column_dimensions['A'].width = 15  # Adjust the width of the date column

    wb.save(file_path)

# Create a new Excel file if it doesn't exist
if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Reason"])  # Add headers
    wb.save(file_path)
    format_excel()  # Apply formatting to the new file

# Route for login page
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username == 'admin' and password == '1234':
            return redirect(url_for('index'))
        else:
            return "Invalid credentials, please try again."
    return render_template('login.html')

# Route to serve the main HTML page
@app.route('/')
def index():
    return render_template('index.html')

# Route to mark present
@app.route('/mark-present')
def mark_present():
    wb = load_workbook(file_path)
    ws = wb.active
    date = datetime.now().strftime('%Y-%m-%d')  # Only date, no time
    ws.append([date, "Present"])
    wb.save(file_path)
    format_excel()
    return "Thanks for submitting today's attendance!"

# Route to mark absent
@app.route('/mark-absent', methods=['POST'])
def mark_absent():
    data = request.get_json()
    reason = data.get('reason', '')
    wb = load_workbook(file_path)
    ws = wb.active
    date = datetime.now().strftime('%Y-%m-%d')  # Only date, no time
    ws.append([date, reason])
    wb.save(file_path)
    format_excel()
    return "Your teacher is marked as absent for today."

# Route to download or view the Excel file
@app.route('/download')
def download_file():
    return send_from_directory(directory='static', path='attendance.xlsx', as_attachment=False)

# Route to reset the attendance file
@app.route('/reset-attendance', methods=['GET'])
def reset_attendance():
    global reset_done

    if reset_done:
        return "The attendance file has already been reset. It cannot be reset again."

    # Delete the existing file if it exists
    if os.path.exists(file_path):
        os.remove(file_path)
    
    # Create a new Excel file
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Reason"])  # Add headers
    wb.save(file_path)
    format_excel()  # Apply formatting to the new file
    
    reset_done = True  # Mark the reset as done
    return "Attendance file has been reset successfully!"

if __name__ == "__main__":
    app.run(debug=True)
